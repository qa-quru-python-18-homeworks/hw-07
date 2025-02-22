import csv
import fitz
from io import BytesIO, TextIOWrapper
from zipfile import ZipFile
from openpyxl import load_workbook


def check_archive_exists(zip_path):
    if not zip_path.exists():
        raise FileNotFoundError("Архив не найден")


def check_file_in_archive_exist(zipf: ZipFile, filename: str):
    if filename not in zipf.namelist():
        raise FileNotFoundError(f"Файл {filename} отсутствует в архиве")


def extract_pdf_text_from_zip(
    zip_path: str, pdf_filename: str, page_number: int
) -> str:
    # Проверка существования архива
    check_archive_exists(zip_path)

    with ZipFile(zip_path) as zipf:
        # Проверка существования файла в архиве
        check_file_in_archive_exist(zipf, pdf_filename)

        # Извлечение файла из архива и создание из него PDF-документа
        with zipf.open(pdf_filename) as pdf_file, fitz.open(
            stream=pdf_file.read()
        ) as pdf_doc:
            try:
                # Извлечение текста на конкретной странице PDF-документа
                return pdf_doc[page_number - 1].get_text()
            except IndexError:
                raise ValueError(
                    f"Страница {page_number} отсутствует в файле {pdf_filename}"
                )


def extract_csv_content_from_zip(zip_path: str, csv_filename: str) -> list[dict]:
    # Проверка существования архива
    check_archive_exists(zip_path)

    with ZipFile(zip_path) as zipf:
        # Проверка существования файла в архива
        check_file_in_archive_exist(zipf, csv_filename)

        # Извлечение файла из архива
        with zipf.open(csv_filename) as csv_file:
            # Создание из него CSV-документа
            with TextIOWrapper(csv_file, encoding="utf-8") as csv_doc:
                # Чтение и возвращение как список словарей
                reader = csv.DictReader(csv_doc)
                return list(reader)


def extract_xlsx_row_from_zip(
    zip_path: str,
    xlsx_filename: str,
    sheet_name: str = None,  # Если имя таблицы не передано, то будет браться активная
    row_number: int = 1,  # Если номер строки не передан, то будет браться первая
):
    # Проверка существования архива
    check_archive_exists(zip_path)

    with ZipFile(zip_path, "r") as zipf:
        # Проверка существования файла в архива
        check_file_in_archive_exist(zipf, xlsx_filename)

        # Извлечение файла из архива
        with zipf.open(xlsx_filename) as xlsx_file:
            # Создание из него XLSX-документа
            xlsx_doc = load_workbook(BytesIO(xlsx_file.read()), data_only=True)

            sheet = xlsx_doc[sheet_name] if sheet_name is not None else xlsx_doc.active

            # Workbook класс не поддерживает контекстный менеджер, поэтому закрывается вручную
            xlsx_doc.close()

            if row_number > sheet.max_row:
                raise ValueError(f"В файле {xlsx_filename} нет строки {row_number}")
            return [cell.value for cell in sheet[row_number]]
